import base64
import os
import tempfile

import pytest
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

from parsers.excel_parser import ExcelParser
from parsers.base_models import ChunkData


@pytest.mark.asyncio
async def test_parse_real_basic_and_image():
    # 准备临时PNG图片（1x1透明像素）
    one_px_png_b64 = (
        b"iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Y2oU5wAAAAASUVORK5CYII="
    )
    png_fd, png_path = tempfile.mkstemp(suffix=".png")
    try:
        with os.fdopen(png_fd, "wb") as f:
            f.write(base64.b64decode(one_px_png_b64))

        # 构建包含图片与两个工作表的真实Excel文件
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        # 表头与数据
        ws1["A1"] = "Header1"
        ws1["B1"] = "Header2"
        ws1["A2"] = "Data1"
        ws1["B2"] = "Data2"
        # 插入图片
        img = XLImage(png_path)
        ws1.add_image(img, "A5")

        # 第二个工作表
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "Single Header"
        ws2["A2"] = "Single Data"

        xlsx_fd, xlsx_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(xlsx_fd)
        wb.save(xlsx_path)

        try:
            parser = ExcelParser()
            result = await parser.parse(xlsx_path)

            assert result.success is True
            # 内容：Sheet1标题、Sheet1图片、Sheet1表格、Sheet2标题、Sheet2表格
            content = result.tables
            assert len(content) == 2

            assert len(result.images) == 1
            assert len(result.texts) == 2

            # 校验顺序与关键字段
            assert result.texts[0].type == "text" and result.texts[0].name == "Sheet1"
            assert result.images[0].type == "image"
            assert result.images[0].name == "#/pictures/0"
            assert result.images[0].content.startswith("data:image/")

            assert result.tables[0].type == "table"
            assert result.texts[1].type == "text" and result.texts[1].name == "Sheet2"
            assert result.tables[1].type == "table"
        finally:
            os.remove(xlsx_path)
    finally:
        os.remove(png_path)


@pytest.mark.asyncio
async def test_parse_real_merged_cells():
    # 构建包含合并单元格的真实Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 合并 A1:B1 并设置值
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws["A1"] = "Merged Header"
    # 填充下一行数据
    ws["A2"] = "Value1"
    ws["B2"] = "Value2"

    xlsx_fd, xlsx_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(xlsx_fd)
    wb.save(xlsx_path)

    try:
        parser = ExcelParser()
        result = await parser.parse(xlsx_path)

        assert result.success is True
        # 结构：标题、表格
        assert len(result.tables) == 1
        assert len(result.texts) == 1

        # 表格在索引1
        table_chunk: ChunkData = result.tables[0]
        assert table_chunk.type == "table"

        payload = table_chunk.content
        assert payload.row_headers == ["Merged Header", "Merged Header"]
        assert payload.data == [["Value1", "Value2"]]
        assert payload.rows == 2
        assert payload.columns == 2
    finally:
        os.remove(xlsx_path)


