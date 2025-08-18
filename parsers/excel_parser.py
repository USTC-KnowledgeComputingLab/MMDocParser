"""
Excel文件解析器模块

该模块提供将Excel文件转换为结构化JSON格式的功能，
包括表格数据提取和图片处理。
"""

import base64
import json
import time
import warnings
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore
from openpyxl.drawing.image import Image  # type: ignore
from openpyxl.workbook.workbook import Workbook  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore

from parsers.document_parser import DocumentData, DocumentParser, ParseResult

# 忽略 openpyxl 的特定警告
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# 类型别名
CellValue = str|int|float|bool|None|datetime|date
TableData = list[list[CellValue]]


@dataclass
class ExcelParseConfig:
    """Excel解析配置类"""
    data_only: bool = True
    keep_vba: bool = False
    default_image_format: str = 'png'
    image_description_placeholder: str = "[待生成]"


class ExcelParseError(Exception):
    """Excel解析异常类"""
    pass


class ExcelParser(DocumentParser):
    """Excel文件解析器类"""

    def __init__(self, config: ExcelParseConfig | None = None):
        """
        初始化Excel解析器
        Args:
            config: 解析配置，如果为None则使用默认配置
        """
        super().__init__()
        self.config: ExcelParseConfig = config or ExcelParseConfig()
        self.image_index: int = 0
        self.supported_formats: list[str] = ['.xlsx', '.xls']

    async def parse(self, excel_path: str) -> ParseResult:
        """
        解析Excel文件并保存结果

        Args:
            excel_path: Excel文件路径
            output_dir: 输出目录路径
        Returns:
            ParseResult: 解析结果对象
        Raises:
            ExcelParseError: 当解析失败时抛出
        """
        start_time = time.time()

        try:
            # 转换Excel到JSON格式
            title, document_data = self._excel_to_json(excel_path)

            # 计算处理时间
            processing_time = time.time() - start_time


            return ParseResult(
                title=title,
                document=document_data,
                processing_time=processing_time,
                success=True
            )

        except Exception as e:
            processing_time = time.time() - start_time
            return ParseResult(
                title="",
                document=[],
                processing_time=processing_time,
                success=False,
                error_message=str(e)
            )

    def can_parse(self, file_path: str) -> bool:
        """
        验证输入文件
        Args:
            file_path: 文件路径
        Returns:
            bool: 是否支持解析
        """
        return any(file_path.lower().endswith(fmt) for fmt in self.supported_formats)

    def _excel_to_json(self, excel_path: str) -> tuple[str, list[DocumentData]]:
        """
        将Excel文件转换为JSON格式
        Args:
            excel_path: Excel文件路径
        Returns:
            DocumentData: 文档数据
        """
        # 获取文件名作为标题（不带扩展名）
        title = Path(excel_path).stem

        # 初始化内容列表和图片列表
        content: list[DocumentData] = []
        self.image_index = 0

        # 加载工作簿
        workbook = self._load_workbook(excel_path)

        # 处理每个工作表
        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]

            # 添加工作表标题
            content.append(DocumentData(
                type="text",
                name=sheet_name,
                content=f"工作表 {sheet_index + 1}: {sheet_name}",
                description="工作表标题"
            ))

            # 处理图片
            sheet_images = self._extract_sheet_images(sheet)
            content.extend(sheet_images)

            # 处理表格数据
            table_content = self._extract_table_data(sheet)
            content.append(DocumentData(
                type="table",
                name="表格",
                content=json.dumps(table_content),
                description="表格"
            ))

        # 添加结束文本
        content.append(DocumentData(
            type="text",
            name="结束文本",
            content="",
            description="结束文本"
        ))

        return title, content

    def _load_workbook(self, excel_path: str) -> Workbook:
        """
        加载Excel工作簿
        Args:
            excel_path: Excel文件路径
        Returns:
            Workbook: 加载的工作簿对象
        """
        return load_workbook(
            excel_path,
            data_only=self.config.data_only,
            keep_vba=self.config.keep_vba
        )

    def _extract_sheet_images(self, sheet: Worksheet) -> list[DocumentData]:
        """
        提取工作表中的图片
        Args:
            sheet: 工作表对象
        Returns:
            List[DocumentData]: 图片信息列表
        """
        sheet_images: list[DocumentData] = []

        images = getattr(sheet, '_images', None)
        if not images or not isinstance(images, (list, tuple)):
            return sheet_images

        # 收集图片信息
        for img_obj in images:
            if not isinstance(img_obj, Image):
                continue

            try:
                image_info = self._process_image_object(img_obj)
                if image_info:
                    sheet_images.append(image_info)
            except Exception as e:
                print(f"处理图片失败: {str(e)}")
                continue

        return sheet_images

    def _process_image_object(self, img_obj: Image) -> DocumentData | None:
        """
        处理单个图片对象
        Args:
            img_obj: 图片对象
        Returns:
            Optional[DocumentData]: 图片信息，处理失败时返回None
        """
        try:
            # 获取图片数据
            img_data = img_obj._data()

            # 获取图片格式
            img_format = self._get_image_format(img_obj)

            # 生成Base64编码
            base64_encoded = base64.b64encode(img_data).decode('utf-8')
            uri = f"data:image/{img_format};base64,{base64_encoded}"

            # 创建图片信息
            image_info = DocumentData(
                type="image",
                name=f"#/pictures/{self.image_index}",
                content=uri,
                description=self.config.image_description_placeholder
            )

            self.image_index += 1
            return image_info

        except Exception as e:
            print(f"处理图片对象失败: {str(e)}")
            return None

    def _get_image_format(self, img_obj: Image) -> str:
        """
        获取图片格式
        Args:
            img_obj: 图片对象
        Returns:
            str: 图片格式
        """
        fmt = getattr(img_obj, 'format', None)
        if isinstance(fmt, str) and fmt:
            img_format: str = fmt.lower()
            # 处理JPEG格式的别名
            if img_format == 'jpeg':
                img_format = 'jpg'
            return img_format
        return self.config.default_image_format

    def _process_cell_value(self, cell_value: Any) -> CellValue:
        """
        预处理单元格值，将datetime对象转换为字符串
        Args:
            cell_value: 原始单元格值
        Returns:
            CellValue: 处理后的单元格值
        """
        if cell_value is None:
            return ""

        # 处理datetime对象，转换为ISO格式字符串
        if isinstance(cell_value, datetime):
            return cell_value.strftime("%Y-%m-%d %H:%M:%S")

        # 处理date对象，转换为日期字符串
        if isinstance(cell_value, date):
            return cell_value.strftime("%Y-%m-%d")

        # 处理其他类型
        if isinstance(cell_value, str|int|float|bool):
            return cell_value

        # 对于其他类型，转换为字符串
        return str(cell_value)

    def _extract_table_data(self, sheet: Worksheet) -> dict[str, Any]:
        """
        提取表格数据
        Args:
            sheet: 工作表对象
        Returns:
            Dict[str, Any]: 表格数据
        """
        # 获取合并单元格信息
        merged_ranges = self._get_merged_cells(sheet)
        merged_map = self._create_merged_cell_map(merged_ranges, sheet)

        # 计算表格维度
        max_row = sheet.max_row
        max_col = sheet.max_column

        # 提取所有数据
        all_rows = self._extract_all_rows(sheet, max_row, max_col, merged_map)

        return {
            "dimensions": {
                "rows": len(all_rows),
                "columns": max_col
            },
            "headers": all_rows[0] if all_rows else [],
            "data": all_rows[1:] if len(all_rows) > 1 else []
        }

    def _get_merged_cells(self, sheet: Worksheet) -> dict[tuple[int, int, int, int], CellValue]:
        """
        获取合并单元格信息
        Args:
            sheet: 工作表对象
        Returns:
            Dict: 合并单元格映射
        """
        merged_ranges = {}
        if sheet.merged_cells:
            for merged_range in sheet.merged_cells.ranges:
                min_row, min_col, max_row, max_col = (
                    merged_range.min_row, merged_range.min_col,
                    merged_range.max_row, merged_range.max_col
                )
                merged_value = sheet.cell(row=min_row, column=min_col).value
                merged_ranges[(min_row, min_col, max_row, max_col)] = merged_value
        return merged_ranges

    def _create_merged_cell_map(self, merged_ranges: dict, sheet: Worksheet) -> dict[tuple[int, int], CellValue]:
        """
        创建合并单元格映射
        Args:
            merged_ranges: 合并单元格范围
            sheet: 工作表对象
        Returns:
            Dict: 合并单元格映射
        """
        merged_map = {}
        for (min_row, min_col, max_row, max_col), value in merged_ranges.items():
            # 预处理合并单元格的值
            processed_value = self._process_cell_value(value)
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    merged_map[(row_idx, col_idx)] = processed_value
        return merged_map

    def _extract_all_rows(self, sheet: Worksheet, max_row: int, max_col: int,
                          merged_map: dict[tuple[int, int], CellValue]) -> TableData:
        """
        提取所有行数据
        Args:
            sheet: 工作表对象
            max_row: 最大行数
            max_col: 最大列数
            merged_map: 合并单元格映射
        Returns:
            TableData: 所有行数据
        """
        all_rows = []
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                # 检查是否是合并单元格
                if (row_idx, col_idx) in merged_map:
                    cell_value = merged_map[(row_idx, col_idx)]
                else:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value

                # 预处理单元格值
                processed_value = self._process_cell_value(cell_value)
                row_data.append(processed_value)
            all_rows.append(row_data)

        return all_rows


    def _save_json(self, data: Any, file_path: Path) -> None:
        """
        保存JSON数据到文件
        Args:
            data: 要保存的数据
            file_path: 文件路径
        """
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
